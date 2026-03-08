from __future__ import annotations

from pathlib import Path
from typing import List
import openpyxl

from app.models import Factura as UIFactura, Cliente, DatosFactura, Concepto
from parser.legacy_excel_parser import ExcelFacturaParser, load_catalogs

_CATALOGS_LOADED = False


def _ensure_catalogs_loaded():
    global _CATALOGS_LOADED
    if not _CATALOGS_LOADED:
        load_catalogs()
        _CATALOGS_LOADED = True


def _split_origen(origen: str):
    origen = (origen or "").strip()
    if "::" in origen:
        a, h = origen.split("::", 1)
        return a.strip(), h.strip()
    return origen, ""


def parse_excel_files(paths: List[str]) -> List[UIFactura]:
    _ensure_catalogs_loaded()
    parser = ExcelFacturaParser()
    out: List[UIFactura] = []

    for p in (paths or []):
        ruta = Path(str(p))

        if ruta.name.startswith("~$"):
            continue

        # Seguimos usando openpyxl SOLAMENTE para asegurar el total correcto
        totales_por_hoja = {}
        try:
            wb = openpyxl.load_workbook(ruta, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                total_encontrado = None
                rows = list(ws.iter_rows(values_only=True))

                for row in reversed(rows):
                    row_vals = [val for val in row if val is not None and str(val).strip() != ""]
                    if not row_vals: continue

                    row_str = " ".join(str(v).upper() for v in row_vals)
                    if any(bad in row_str for bad in ["SUBTOTAL", "SUB TOTAL", "LETRA"]):
                        continue

                    if "TOTAL" in row_str and total_encontrado is None:
                        nums = []
                        for val in row_vals:
                            try:
                                clean_val = str(val).replace("$", "").replace(",", "").strip()
                                nums.append(float(clean_val))
                            except ValueError:
                                pass
                        if nums:
                            total_encontrado = nums[-1]

                if total_encontrado is not None:
                    totales_por_hoja[sheet_name] = total_encontrado
            wb.close()
        except Exception:
            pass

        # Llamamos al Legacy Parser (ahora operado y curado)
        legacy_facturas = parser.parse_file(ruta)

        for lf in legacy_facturas:
            origen_str = getattr(lf.archivo, "name", ruta.name) if lf.archivo else ruta.name
            archivo_origen, hoja_origen = _split_origen(origen_str)
            proveedor_bd = (lf.proveedor or "").strip()

            cliente = Cliente(rfc=(lf.rfc or None), proveedor=proveedor_bd)

            datos = DatosFactura(
                uso_cfdi=(lf.uso_cfdi or None),
                metodo_pago=(lf.metodo_pago or None),
                forma_pago=(lf.forma_pago or None),
                es_usd=bool(getattr(lf, "es_usd", False)),
                tipo_cambio="",
                info_extra=getattr(lf, "info_extra", ""),  # ¡Pasamos la nota de la OBRA!
            )

            conceptos_ui: List[Concepto] = []
            for c in (lf.conceptos or []):
                cantidad = float(c.cantidad or 0.0) if c.cantidad is not None else 0.0
                precio_unit = float(c.precio_unitario or 0.0) if c.precio_unitario is not None else 0.0
                importe = float(c.importe) if c.importe is not None else None

                conceptos_ui.append(
                    Concepto(
                        cantidad=cantidad,
                        clave_unidad=str(c.clave_unidad or "").strip(),
                        clave_prod_serv=str(c.clave_prod_serv or "").strip(),
                        concepto=str(c.descripcion or "").strip(),
                        precio_unitario=precio_unit,
                        importe=importe,
                    )
                )

            total_inteligente = None
            if hoja_origen and hoja_origen in totales_por_hoja:
                total_inteligente = totales_por_hoja[hoja_origen]
            elif totales_por_hoja:
                total_inteligente = max(totales_por_hoja.values())

            if total_inteligente is not None and total_inteligente > 0:
                total_val = float(total_inteligente)
            else:
                total_val = float(lf.total or 0.0) if lf.total is not None else 0.0

            factura_id = f"{archivo_origen}::{hoja_origen}" if hoja_origen else archivo_origen

            out.append(
                UIFactura(
                    id=factura_id,
                    cliente=cliente,
                    datos_factura=datos,
                    conceptos=conceptos_ui,
                    total=total_val,
                    archivo_origen=archivo_origen,
                    hoja_origen=hoja_origen,
                )
            )

    return out