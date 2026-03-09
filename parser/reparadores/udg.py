# parser/reparadores/udg.py
from pathlib import Path
import openpyxl


def reparar(ruta_str: str, perfil: str, *args) -> tuple[list[str], int]:
    errores = 0
    archivos_corregidos = []
    try:
        ruta = Path(ruta_str)
        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb.active

        header_idx = None
        col_map = {}

        # Encontrar los encabezados "creativos"
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_str = " ".join(str(c).upper() for c in row if c)
            if "DESCRIPCIÓN" in row_str and "CANTIDAD SOLICITADA" in row_str:
                header_idx = idx
                for c_idx, h in enumerate(row):
                    val = str(h).upper() if h else ""
                    if "DESCRIPCIÓN" in val:
                        col_map['desc'] = c_idx
                    elif "UNIDAD" in val:
                        col_map['uni_nombre'] = c_idx
                    elif "CANTIDAD SOLICITADA" in val:
                        col_map['cant'] = c_idx
                    elif "EXISTENCIA" in val:
                        col_map['pu'] = c_idx
                    elif "OBSERVACIONES" in val:
                        col_map['imp'] = c_idx
                break

        if header_idx is not None:
            molde_wb = openpyxl.Workbook()
            m_ws = molde_wb.active
            m_ws.title = "FACT 1"

            # Inyectamos cliente genérico de UDG
            m_ws.cell(row=2, column=1, value="PROVEEDOR")
            m_ws.cell(row=2, column=2, value="REKLAMSA")
            m_ws.cell(row=4, column=1, value="RAZON SOCIAL")
            m_ws.cell(row=4, column=2, value="UNIVERSIDAD DE GUADALAJARA")
            m_ws.cell(row=5, column=1, value="RFC:")
            m_ws.cell(row=5, column=2, value="UGU250907MH5")

            m_ws.cell(row=12, column=1, value="Uso de CFDI:")
            m_ws.cell(row=12, column=3, value="G03")
            m_ws.cell(row=13, column=1, value="Forma de pago:")
            m_ws.cell(row=13, column=3, value="99")
            m_ws.cell(row=15, column=1, value="Método de Pago:")
            m_ws.cell(row=15, column=3, value="PPD")

            headers = ["CANTIDAD", "CLAVE UNIDAD SAT", "DESCRIPCION UNIDAD SAT", "CLAVE PRODUCTO O SERVICIO SAT",
                       "DESCRIP PROD/SERV SAT", "CONCEPTO", "Precio Unitario", "Descuentos", "Impuesto SAT", "Impuesto",
                       "Importe"]
            for j, h in enumerate(headers, 1): m_ws.cell(row=19, column=j, value=h)

            row_out = 20
            subtotal_calc = 0.0

            for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
                desc = str(row[col_map.get('desc', -1)] or "").strip()

                # --- ESCUDO 1: PARAR AL LLEGAR A LAS FIRMAS ---
                if "FIRMAS" in desc.upper() or "SOLICITANTE" in desc.upper():
                    break

                if not desc or desc.upper() == "NONE": continue

                cant_raw = row[col_map.get('cant', -1)]
                if not cant_raw: continue

                # --- ESCUDO 2: BLINDAJE DE NÚMEROS ---
                try:
                    cant = float(cant_raw)
                except ValueError:
                    continue  # Si alguien escribió letras en la cantidad, saltamos la fila

                uni_nombre = str(row[col_map.get('uni_nombre', -1)] or "").strip().lower()
                pu = row[col_map.get('pu', -1)]
                imp = row[col_map.get('imp', -1)]

                try:
                    pu = float(pu)
                except:
                    pu = 0.0
                try:
                    imp = float(imp)
                except:
                    imp = cant * pu

                if pu == 0.0 and imp == 0.0: continue

                subtotal_calc += imp

                # Mapeo de unidades manual a claves SAT
                clave_unidad = "H87"  # Pieza por defecto
                if "caja" in uni_nombre:
                    clave_unidad = "XBX"
                elif "paquete" in uni_nombre:
                    clave_unidad = "XPK"
                elif "rollo" in uni_nombre:
                    clave_unidad = "XRO"
                elif "servicio" in uni_nombre:
                    clave_unidad = "E48"

                m_ws.cell(row=row_out, column=1, value=cant)
                m_ws.cell(row=row_out, column=2, value=clave_unidad)
                m_ws.cell(row=row_out, column=4, value="01010101")  # Clave genérica para evitar que el parser truene
                m_ws.cell(row=row_out, column=6, value=desc)
                m_ws.cell(row=row_out, column=7, value=pu)
                m_ws.cell(row=row_out, column=11, value=imp)
                row_out += 1

            m_ws.cell(row=row_out + 2, column=2, value="SUBTOTAL")
            m_ws.cell(row=row_out + 2, column=3, value=subtotal_calc)
            m_ws.cell(row=row_out + 3, column=2, value="IVA")
            m_ws.cell(row=row_out + 3, column=3, value=subtotal_calc * 0.16)
            m_ws.cell(row=row_out + 4, column=2, value="TOTAL")
            m_ws.cell(row=row_out + 4, column=3, value=subtotal_calc * 1.16)

            base_name = ruta.stem
            nueva_ruta = ruta.parent / f"[CORREGIDO]_{base_name}.xlsx"
            molde_wb.save(nueva_ruta)
            archivos_corregidos.append(str(nueva_ruta))
        wb.close()
    except Exception as e:
        print(f"Error reparando UDG: {e}")
        errores += 1

    return archivos_corregidos, errores
