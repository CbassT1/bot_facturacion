# parser/reparadores/grajeda.py
import re
from pathlib import Path
import openpyxl


def reparar(ruta_str: str, perfil: str, *args) -> tuple[list[str], int]:
    errores = 0
    archivos_corregidos = []
    try:
        ruta = Path(ruta_str)
        wb = openpyxl.load_workbook(ruta, data_only=True)

        # Buscar la hoja "Solicitud 3.3"
        hoja_objetivo = None
        for sn in wb.sheetnames:
            if "SOLICITUD 3.3" in sn.upper() or "SOLICITUD" in sn.upper():
                hoja_objetivo = wb[sn]
                break

        if not hoja_objetivo:
            hoja_objetivo = wb.active

        # Variables para rescatar
        prov = razon_social = rfc = clave_sat = ""
        uso = "G03"
        forma = "99"
        metodo = "PPD"

        conceptos_raw = []
        leyendo_conceptos = False

        for row in hoja_objetivo.iter_rows(values_only=True):
            row_str = " ".join(str(c).upper() for c in row if c)

            # --- CAZADORES DE METADATOS ---
            if "FACTURADOR" in row_str:
                for i, c in enumerate(row):
                    if c and "FACTURADOR" in str(c).upper() and i + 1 < len(row):
                        prov = str(row[i + 1]).strip()
                        break
            if "CLIENTE" in row_str:
                for i, c in enumerate(row):
                    if c and "CLIENTE" in str(c).upper() and i + 1 < len(row):
                        razon_social = str(row[i + 1]).strip()
                    # A veces RFC está en la misma fila
                    if c and "R.F.C" in str(c).upper().replace(" ", "") and i + 1 < len(row):
                        rfc = str(row[i + 1]).strip()

            if "CLAVE PROD" in row_str:
                for i, c in enumerate(row):
                    if c and "CLAVE PROD" in str(c).upper():
                        for j in range(i + 1, len(row)):
                            if row[j] and str(row[j]).strip() != "":
                                clave_sat = str(row[j]).strip()
                                break
                        break

            # CORRECCIÓN: Búsqueda persistente hacia la derecha saltando celdas vacías
            if "USO CFDI" in row_str:
                for i, c in enumerate(row):
                    if c and "USO CFDI" in str(c).upper():
                        for j in range(i + 1, len(row)):
                            val = str(row[j]).strip() if row[j] else ""
                            if val:
                                uso = val[:3]
                                break  # Solo se detiene cuando ya encontró el dato
                        break

            if "METODO DE PAGO" in row_str:
                for i, c in enumerate(row):
                    if c and "METODO DE PAGO" in str(c).upper():
                        for j in range(i + 1, len(row)):
                            val = str(row[j]).strip() if row[j] else ""
                            if val:
                                metodo = val[:3]
                                break
                        break

            if "FORMA DE PAGO" in row_str:
                for i, c in enumerate(row):
                    if c and "FORMA DE PAGO" in str(c).upper():
                        for j in range(i + 1, len(row)):
                            val = str(row[j]).strip() if row[j] else ""
                            if val:
                                m = re.search(r'\d{2}', val)
                                if m:
                                    forma = m.group(0)
                                break
                        break

            # --- CAZADOR DE LA TABLA DE CONCEPTOS ---
            if "CANTIDAD" in row_str and "UNIDAD" in row_str and "PRECIO" in row_str:
                leyendo_conceptos = True
                continue

            if leyendo_conceptos:
                if not any(row):
                    leyendo_conceptos = False
                    continue

                c_idx = -1
                for i, c in enumerate(row):
                    if c is not None and str(c).strip() != "":
                        c_idx = i
                        break

                if c_idx != -1 and len(row) > c_idx + 3:
                    cant = row[c_idx]
                    uni = str(row[c_idx + 1] or "E48").strip()[:3]
                    desc = str(row[c_idx + 2] or "").strip()
                    pu = row[c_idx + 3]

                    if cant and desc:
                        conceptos_raw.append({
                            'cant': cant, 'uni': uni, 'desc': desc, 'pu': pu
                        })

        # ==========================================
        # CONSTRUIR EL MOLDE DE FORMATO VACÍO
        # ==========================================
        molde_wb = openpyxl.Workbook()
        m_ws = molde_wb.active
        m_ws.title = "FACT 1"

        m_ws.cell(row=2, column=1, value="PROVEEDOR")
        m_ws.cell(row=2, column=2, value=prov)
        m_ws.cell(row=4, column=1, value="RAZON SOCIAL")
        m_ws.cell(row=4, column=2, value=razon_social)
        m_ws.cell(row=5, column=1, value="RFC:")
        m_ws.cell(row=5, column=2, value=rfc)

        m_ws.cell(row=12, column=1, value="Uso de CFDI:")
        m_ws.cell(row=12, column=3, value=uso)
        m_ws.cell(row=13, column=1, value="Forma de pago:")
        m_ws.cell(row=13, column=3, value=forma)
        m_ws.cell(row=15, column=1, value="Método de Pago:")
        m_ws.cell(row=15, column=3, value=metodo)

        headers = ["CANTIDAD", "CLAVE UNIDAD SAT", "DESCRIPCION UNIDAD SAT", "CLAVE PRODUCTO O SERVICIO SAT",
                   "DESCRIP PROD/SERV SAT", "CONCEPTO", "Precio Unitario", "Descuentos", "Impuesto SAT", "Impuesto",
                   "Importe"]
        for j, h in enumerate(headers, 1): m_ws.cell(row=19, column=j, value=h)

        row_out = 20
        subtotal_calc = 0.0

        for c in conceptos_raw:
            try:
                p = float(c['pu'])
            except:
                p = 0.0
            try:
                ca = float(c['cant'])
            except:
                ca = 0.0
            imp = p * ca
            subtotal_calc += imp

            m_ws.cell(row=row_out, column=1, value=ca)
            m_ws.cell(row=row_out, column=2, value=c['uni'])
            m_ws.cell(row=row_out, column=4, value=clave_sat)
            m_ws.cell(row=row_out, column=6, value=c['desc'])
            m_ws.cell(row=row_out, column=7, value=p)
            m_ws.cell(row=row_out, column=11, value=imp)
            row_out += 1

        m_ws.cell(row=row_out + 2, column=2, value="SUBTOTAL")
        m_ws.cell(row=row_out + 2, column=3, value=subtotal_calc)
        m_ws.cell(row=row_out + 3, column=2, value="IVA")
        m_ws.cell(row=row_out + 3, column=3, value=subtotal_calc * 0.16)
        m_ws.cell(row=row_out + 4, column=2, value="TOTAL")
        m_ws.cell(row=row_out + 4, column=3, value=subtotal_calc * 1.16)

        base_name = ruta.stem
        nueva_ruta = ruta.parent / f"[CORREGIDO]_{base_name}.xlsx"
        molde_wb.save(nueva_ruta)
        archivos_corregidos.append(str(nueva_ruta))
        wb.close()

    except Exception as e:
        print(f"Error reparando Solicitud Grajeda: {e}")
        errores += 1

    return archivos_corregidos, errores
