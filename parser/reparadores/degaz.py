from pathlib import Path
import openpyxl

def reparar(ruta_str: str, perfil: str, *args) -> tuple[list[str], int]:
    errores = 0
    archivos_corregidos = []
    try:
        ruta = Path(ruta_str)
        wb = openpyxl.load_workbook(ruta, data_only=True)
        hoja_pedido = hoja_orden = None
        for sn in wb.sheetnames:
            if "ORDEN" in sn.upper(): hoja_orden = wb[sn]
            if "PEDIDO" in sn.upper(): hoja_pedido = wb[sn]

        if hoja_pedido and hoja_orden:
            razon_social = rfc = ""
            for row in hoja_pedido.iter_rows(min_row=1, max_row=10, values_only=True):
                row_str = " ".join(str(c).upper() for c in row if c)
                if "RAZON SOCIAL" in row_str: razon_social = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                elif "RFC" in row_str: rfc = str(row[1]).strip() if len(row) > 1 and row[1] else ""

            header_idx = None
            col_map = {}
            for idx, row in enumerate(hoja_pedido.iter_rows(values_only=True), 1):
                row_str = " ".join(str(c).upper() for c in row if c)
                if "PRECIO UNITARIO" in row_str and "CONCEPTO" in row_str:
                    header_idx = idx
                    for c_idx, h in enumerate(row):
                        val = str(h).upper() if h else ""
                        if "PROVEEDOR" in val or "EMPRESA" in val: col_map['prov'] = c_idx
                        elif "USO DEL" in val: col_map['uso'] = c_idx
                        elif "METODO" in val: col_map['metodo'] = c_idx
                        elif "FORMA" in val: col_map['forma'] = c_idx
                        elif "UNIDAD MEDIDA" in val or "CLAVE" in val and "PRODUCTO" not in val: col_map['uni'] = c_idx
                        elif "CANTIDAD" in val: col_map['cant'] = c_idx
                        elif "PRECIO UNITARIO" in val: col_map['pu'] = c_idx
                        elif "SUBTOTAL" in val: col_map['sub'] = c_idx
                        elif "TOTAL" in val and "SUB" not in val: col_map['tot'] = c_idx
                        elif "PRODUCTO" in val: col_map['prod'] = c_idx
                        elif "CONCEPTO" in val or "DESCRIPCION" in val: col_map['desc'] = c_idx
                    break

            facturas_generadas = 0
            for row in hoja_pedido.iter_rows(min_row=header_idx + 1, values_only=True):
                if not row[col_map.get('cant', 0)] or not row[col_map.get('desc', 0)]: continue
                cant = row[col_map.get('cant')]
                if str(cant).strip() == "": continue

                prov = row[col_map.get('prov')]
                uso = row[col_map.get('uso')]
                metodo = row[col_map.get('metodo')]
                forma = row[col_map.get('forma')]
                uni = row[col_map.get('uni')]
                pu = row[col_map.get('pu')]
                sub = row[col_map.get('sub')]
                tot = row[col_map.get('tot')]
                prod = row[col_map.get('prod')]
                desc = str(row[col_map.get('desc')]).strip()

                extra_desc = po_number = ""
                found_r = found_c = -1
                for r_idx, o_row in enumerate(hoja_orden.iter_rows(values_only=True), 1):
                    for c_idx, o_val in enumerate(o_row, 1):
                        if isinstance(o_val, str) and desc.upper() in o_val.upper():
                            found_r, found_c = r_idx, c_idx
                            break
                    if found_r != -1: break

                if found_r != -1:
                    below = hoja_orden.cell(row=found_r + 1, column=found_c).value
                    if below and str(below).strip() and "DESCRIPCIÓN" not in str(below).upper():
                        extra_desc = str(below).strip()
                    for sr in range(found_r, 0, -1):
                        for sc in range(max(1, found_c - 4), found_c + 5):
                            c_val = hoja_orden.cell(row=sr, column=sc).value
                            if c_val and "COTIZACIÓN" in str(c_val).upper():
                                po_val = hoja_orden.cell(row=sr, column=sc + 1).value
                                if po_val: po_number = str(po_val).strip()
                                break
                        if po_number: break

                desc_completa = f"{desc}\n{extra_desc}" if extra_desc else desc

                facturas_generadas += 1
                molde_wb = openpyxl.Workbook()
                m_ws = molde_wb.active
                m_ws.title = "FACT 1"
                m_ws.cell(row=2, column=1, value="PROVEEDOR")
                m_ws.cell(row=2, column=2, value=prov)
                m_ws.cell(row=4, column=1, value="RAZON SOCIAL")
                m_ws.cell(row=4, column=2, value=razon_social)
                m_ws.cell(row=5, column=1, value="RFC:")
                m_ws.cell(row=5, column=2, value=rfc)
                m_ws.cell(row=12, column=1, value="Uso de CFDI:")
                m_ws.cell(row=12, column=3, value=uso)
                m_ws.cell(row=13, column=1, value="Forma de pago:")
                m_ws.cell(row=13, column=3, value=forma)
                m_ws.cell(row=15, column=1, value="Método de Pago:")
                m_ws.cell(row=15, column=3, value=metodo)

                if po_number: m_ws.cell(row=18, column=1, value=f"OBRA: No. Cotización: {po_number}")

                headers = ["CANTIDAD", "CLAVE UNIDAD SAT", "DESCRIPCION UNIDAD SAT", "CLAVE PRODUCTO O SERVICIO SAT", "DESCRIP PROD/SERV SAT", "CONCEPTO", "Precio Unitario", "Descuentos", "Impuesto SAT", "Impuesto", "Importe"]
                for j, h in enumerate(headers, 1): m_ws.cell(row=19, column=j, value=h)

                m_ws.cell(row=20, column=1, value=cant)
                m_ws.cell(row=20, column=2, value=uni)
                m_ws.cell(row=20, column=4, value=prod)
                m_ws.cell(row=20, column=6, value=desc_completa)
                m_ws.cell(row=20, column=7, value=pu)
                m_ws.cell(row=20, column=11, value=sub)
                m_ws.cell(row=23, column=2, value="SUBTOTAL")
                m_ws.cell(row=23, column=3, value=sub)
                m_ws.cell(row=24, column=2, value="IVA")
                iva_calc = tot - sub if tot and sub else 0
                m_ws.cell(row=24, column=3, value=iva_calc)
                m_ws.cell(row=25, column=2, value="TOTAL")
                m_ws.cell(row=25, column=3, value=tot)

                base_name = ruta.stem
                nueva_ruta = ruta.parent / f"[CORREGIDO]_{base_name}_Factura_{facturas_generadas}.xlsx"
                molde_wb.save(nueva_ruta)
                archivos_corregidos.append(str(nueva_ruta))
        wb.close()
    except Exception as e:
        print(f"Error DEGAZ {ruta_str}: {e}")
        errores += 1
    return archivos_corregidos, errores
