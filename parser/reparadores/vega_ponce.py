from pathlib import Path
import openpyxl


def reparar(ruta_str: str, perfil: str, *args) -> tuple[list[str], int]:
    errores = 0
    archivos_corregidos = []
    try:
        ruta = Path(ruta_str)
        wb = openpyxl.load_workbook(ruta, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            obra_texto = ""
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and str(cell.value).strip().upper().startswith("OBRA:"):
                        obra_texto = str(cell.value).strip()
                        cell.value = ""

            header_row_idx = col_unidad = col_producto = None
            for idx, row in enumerate(ws.iter_rows(), start=1):
                row_str = " ".join(str(c.value).upper() for c in row if c.value)
                if "CANTIDAD" in row_str and "UNIDAD" in row_str:
                    header_row_idx = idx
                    for c_idx, cell in enumerate(row, start=1):
                        val_str = str(cell.value).strip().upper() if cell.value else ""
                        if "CLAVE UNIDAD" in val_str:
                            col_unidad = c_idx
                        elif "CLAVE PROD" in val_str or "PRODUCTO O SERVICIO" in val_str:
                            col_producto = c_idx
                    break

            if header_row_idx and col_unidad and col_producto:
                for r in range(header_row_idx + 1, ws.max_row + 1):
                    cell_uni = ws.cell(row=r, column=col_unidad)
                    cell_prod = ws.cell(row=r, column=col_producto)
                    val_u = str(cell_uni.value).strip().replace(".0", "") if cell_uni.value else ""
                    if val_u.isdigit() and len(val_u) >= 6:
                        cell_prod.value = val_u
                        cell_uni.value = "H87"
                if obra_texto:
                    ws.insert_rows(header_row_idx)
                    ws.cell(row=header_row_idx, column=1).value = obra_texto

        nueva_ruta = ruta.parent / f"[CORREGIDO]_{ruta.name}"
        wb.save(nueva_ruta)
        wb.close()
        archivos_corregidos.append(str(nueva_ruta))
    except Exception as e:
        print(f"Error Vega Ponce {ruta_str}: {e}")
        errores += 1
    return archivos_corregidos, errores
