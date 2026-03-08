from pathlib import Path
import openpyxl


def reparar(ruta_str: str, perfil: str, cliente_sel: str) -> tuple[list[str], int]:
    errores = 0
    archivos_corregidos = []
    try:
        ruta = Path(ruta_str)
        wb = openpyxl.load_workbook(ruta, data_only=True)
        razon_social = rfc = ""
        if "Campanas" in cliente_sel:
            razon_social, rfc = "Las Campanas", "SAC1212284C7"
        elif "Escobedo" in cliente_sel:
            razon_social, rfc = "AG Escobedo", "SAE2107264V8"
        elif "Ancira" in cliente_sel:
            razon_social, rfc = "AG Ancira", "SAG950408LE8"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_idx = None
            col_map = {}
            for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_str = " ".join(str(c).upper() for c in row if c)
                if "PROVEEDOR" in row_str and "DESCRIPCIÓN" in row_str and "CLAVE SAT" in row_str:
                    header_idx = idx
                    for c_idx, h in enumerate(row):
                        val = str(h).upper() if h else ""
                        if "PROVEEDOR" in val:
                            col_map['prov'] = c_idx
                        elif "CANTIDAD" in val:
                            col_map['cant'] = c_idx
                        elif "DESCRIPCI" in val:
                            col_map['desc'] = c_idx
                        elif "CLAVE SAT" in val:
                            col_map['prod'] = c_idx
                        elif "IMPORTE" in val or "PRECIO" in val:
                            col_map['pu'] = c_idx
                        elif "SUB" in val:
                            col_map['sub'] = c_idx
                        elif "TOTAL" in val and "SUB" not in val:
                            col_map['tot'] = c_idx
                        elif "USO" in val:
                            col_map['uso'] = c_idx
                    break

            if header_idx is not None:
                facturas_generadas = 0
                for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
                    prov = row[col_map.get('prov', -1)] if col_map.get('prov') is not None else None
                    desc = row[col_map.get('desc', -1)] if col_map.get('desc') is not None else None

                    if not prov or not desc or str(prov).strip() == "" or str(desc).strip().upper() in ["DESCRIPCIÓN",
                                                                                                        "CONCEPTO"]: continue

                    cant = row[col_map.get('cant', -1)] if col_map.get('cant') is not None else 1
                    prod = row[col_map.get('prod', -1)] if col_map.get('prod') is not None else ""
                    pu = row[col_map.get('pu', -1)] if col_map.get('pu') is not None else 0
                    sub = row[col_map.get('sub', -1)] if col_map.get('sub') is not None else pu
                    tot = row[col_map.get('tot', -1)] if col_map.get('tot') is not None else sub
                    uso = row[col_map.get('uso', -1)] if col_map.get('uso') is not None else "G03"

                    facturas_generadas += 1
                    molde_wb = openpyxl.Workbook()
                    m_ws = molde_wb.active
                    m_ws.title = "FACT 1"
                    m_ws.cell(row=2, column=1, value="PROVEEDOR")
                    m_ws.cell(row=2, column=2, value=str(prov).strip())
                    m_ws.cell(row=4, column=1, value="RAZON SOCIAL")
                    m_ws.cell(row=4, column=2, value=razon_social)
                    m_ws.cell(row=5, column=1, value="RFC:")
                    m_ws.cell(row=5, column=2, value=rfc)
                    m_ws.cell(row=12, column=1, value="Uso de CFDI:")
                    m_ws.cell(row=12, column=3, value=str(uso).strip())
                    m_ws.cell(row=13, column=1, value="Forma de pago:")
                    m_ws.cell(row=13, column=3, value="99")
                    m_ws.cell(row=15, column=1, value="Método de Pago:")
                    m_ws.cell(row=15, column=3, value="PPD")

                    headers = ["CANTIDAD", "CLAVE UNIDAD SAT", "DESCRIPCION UNIDAD SAT",
                               "CLAVE PRODUCTO O SERVICIO SAT", "DESCRIP PROD/SERV SAT", "CONCEPTO", "Precio Unitario",
                               "Descuentos", "Impuesto SAT", "Impuesto", "Importe"]
                    for j, h in enumerate(headers, 1): m_ws.cell(row=19, column=j, value=h)

                    m_ws.cell(row=20, column=1, value=cant)
                    m_ws.cell(row=20, column=2, value="E48")
                    m_ws.cell(row=20, column=4, value=str(prod).strip().replace(".0", ""))
                    m_ws.cell(row=20, column=6, value=str(desc).strip())
                    m_ws.cell(row=20, column=7, value=pu)
                    m_ws.cell(row=20, column=11, value=sub)
                    m_ws.cell(row=23, column=2, value="SUBTOTAL")
                    m_ws.cell(row=23, column=3, value=sub)
                    m_ws.cell(row=24, column=2, value="IVA")
                    try:
                        iva_calc = float(tot) - float(sub)
                    except:
                        iva_calc = 0
                    m_ws.cell(row=24, column=3, value=iva_calc)
                    m_ws.cell(row=25, column=2, value="TOTAL")
                    m_ws.cell(row=25, column=3, value=tot)

                    base_name = ruta.stem
                    nueva_ruta = ruta.parent / f"[CORREGIDO]_{base_name}_Fila_{facturas_generadas}.xlsx"
                    molde_wb.save(nueva_ruta)
                    archivos_corregidos.append(str(nueva_ruta))
        wb.close()
    except Exception as e:
        print(f"Error Gasolinera {ruta_str}: {e}")
        errores += 1
    return archivos_corregidos, errores
