# parser/reparadores/diegza.py
import re
from pathlib import Path
import openpyxl


def _limpiar_concepto(raw_desc: str, clave_prod: str) -> str:
    """
    Toma el concepto sucio y extrae únicamente la descripción real del producto.
    """
    # 1. Quitar la palabra "PRODUCTO" inicial
    desc = re.sub(r'(?i)^\s*PRODUCTO\s*', '', raw_desc)

    # 2. Quitar la clave numérica y el guion inicial
    if clave_prod:
        desc = desc.replace(clave_prod, '')
    desc = re.sub(r'^\s*-\s*', '', desc).strip()

    # 3. Si hay un salto de línea (Enter), es seguro que separa la clave del producto
    if '\n' in desc:
        partes = desc.split('\n', 1)
        desc = partes[1].strip()
    else:
        # 4. Si NO hay salto de línea (como en "general PORRON")
        # Buscamos la transición: La última minúscula seguida de un espacio y luego MAYÚSCULAS.
        match = re.search(r'[a-záéíóúñ\)]\s+([A-ZÁÉÍÓÚÑ0-9]{2,}.*)', desc)
        if match:
            desc = match.group(1).strip()

    # Limpieza final por si quedó basura
    desc = re.sub(r'^\s*-\s*', '', desc).strip()
    return desc


def reparar_diegza(ruta_str: str, cliente_sel: str, *args) -> tuple[list[str], int]:
    errores = 0
    archivos_corregidos = []

    try:
        ruta = Path(ruta_str)
        wb = openpyxl.load_workbook(ruta, data_only=True)

        razon_social = "DIEGZA" if "DIEGZA" in cliente_sel else "CLINNSA"
        rfc = "DIE100730969" if "DIEGZA" in cliente_sel else "CLI180507CN5"
        proveedor_default = "BETANSA"

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 1. Buscar las columnas
            header_idx = None
            col_map = {}
            for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_str = " ".join(str(c).upper() for c in row if c)
                if "CANTIDAD" in row_str and "CONCEPTO" in row_str and "PRECIO" in row_str:
                    header_idx = idx
                    for c_idx, h in enumerate(row):
                        val = str(h).upper() if h else ""
                        if "CANTIDAD" in val:
                            col_map['cant'] = c_idx
                        elif "CONCEPTO" in val:
                            col_map['desc'] = c_idx
                        elif "PRECIO" in val:
                            col_map['pu'] = c_idx
                        elif "IMPORTE" in val:
                            col_map['imp'] = c_idx
                    break

            if header_idx is not None:
                # 2. Crear Molde
                molde_wb = openpyxl.Workbook()
                m_ws = molde_wb.active
                m_ws.title = "FACT 1"

                m_ws.cell(row=2, column=1, value="PROVEEDOR")
                m_ws.cell(row=2, column=2, value=proveedor_default)
                m_ws.cell(row=4, column=1, value="RAZON SOCIAL")
                m_ws.cell(row=4, column=2, value=razon_social)
                m_ws.cell(row=5, column=1, value="RFC:")
                m_ws.cell(row=5, column=2, value=rfc)

                m_ws.cell(row=12, column=1, value="Uso de CFDI:")
                m_ws.cell(row=12, column=3, value="G03")
                m_ws.cell(row=13, column=1, value="Forma de pago:")
                m_ws.cell(row=13, column=3, value="99")
                m_ws.cell(row=15, column=1, value="Método de Pago:")
                m_ws.cell(row=15, column=3, value="PPD")

                headers = ["CANTIDAD", "CLAVE UNIDAD SAT", "DESCRIPCION UNIDAD SAT", "CLAVE PRODUCTO O SERVICIO SAT",
                           "DESCRIP PROD/SERV SAT", "CONCEPTO", "Precio Unitario", "Descuentos", "Impuesto SAT",
                           "Impuesto", "Importe"]
                for j, h in enumerate(headers, 1): m_ws.cell(row=19, column=j, value=h)

                row_out = 20
                subtotal_calc = 0.0

                # 3. Extraer y procesar cada fila
                for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
                    if not row[col_map.get('cant', 0)] or not row[col_map.get('desc', 0)]: continue

                    cant = row[col_map.get('cant')]
                    if str(cant).strip() == "": continue

                    raw_desc = str(row[col_map.get('desc', -1)]).strip()
                    pu = row[col_map.get('pu', -1)]
                    try:
                        pu = float(pu)
                    except:
                        pu = 0.0

                    imp = row[col_map.get('imp', -1)]
                    try:
                        imp = float(imp)
                    except:
                        imp = float(cant) * pu

                    subtotal_calc += imp

                    # Atrapar la clave SAT de 8 dígitos
                    clave_prod = "01010101"
                    match = re.search(r'\b\d{8}\b', raw_desc)
                    if match:
                        clave_prod = match.group(0)

                    # Llamamos al limpiador avanzado
                    clean_desc = _limpiar_concepto(raw_desc, clave_prod)

                    # Vaciar en el molde
                    m_ws.cell(row=row_out, column=1, value=cant)
                    m_ws.cell(row=row_out, column=2, value="H87")
                    m_ws.cell(row=row_out, column=4, value=clave_prod)
                    m_ws.cell(row=row_out, column=6, value=clean_desc)
                    m_ws.cell(row=row_out, column=7, value=pu)
                    m_ws.cell(row=row_out, column=11, value=imp)
                    row_out += 1

                # Escribir los totales
                m_ws.cell(row=row_out + 2, column=2, value="SUBTOTAL")
                m_ws.cell(row=row_out + 2, column=3, value=subtotal_calc)
                m_ws.cell(row=row_out + 3, column=2, value="IVA")
                m_ws.cell(row=row_out + 3, column=3, value=subtotal_calc * 0.16)
                m_ws.cell(row=row_out + 4, column=2, value="TOTAL")
                m_ws.cell(row=row_out + 4, column=3, value=subtotal_calc * 1.16)

                # Guardar cada hoja
                base_name = ruta.stem
                nueva_ruta = ruta.parent / f"[CORREGIDO]_{base_name}_{sheet_name}.xlsx"
                molde_wb.save(nueva_ruta)
                archivos_corregidos.append(str(nueva_ruta))

        wb.close()
    except Exception as e:
        print(f"Error reparando {ruta_str}: {e}")
        errores += 1

    return archivos_corregidos, errores
