# app/ui/frames/reportes.py
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
from datetime import datetime, timedelta

# Importaciones protegidas
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment

    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from tkcalendar import DateEntry

    HAS_CALENDAR = True
except ImportError:
    HAS_CALENDAR = False

from app.database.database import SessionLocal, FacturaGuardada
from app.ui.theme import get_pal


class ReportesFrame(ttk.Frame):
    def __init__(self, master, controller):
        super().__init__(master)
        self.controller = controller
        self._build_ui()

    def _build_ui(self):
        pal = get_pal(self.controller)

        # --- Cabecera ---
        header = ttk.Frame(self)
        header.pack(fill="x", padx=12, pady=(12, 6))

        ttk.Button(header, text="← Menú Principal", command=lambda: self.controller.show("menu")).pack(side="left")
        ttk.Label(header, text="Reportes y Exportación", font=("Segoe UI", 16, "bold")).pack(side="left", padx=(12, 0))
        ttk.Button(header, text=self.controller.theme_button_label(), command=self._toggle_theme).pack(side="right")

        # --- Contenedor Central ---
        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=16, pady=20)

        card = ttk.Frame(body, style="Card.TFrame")
        card.place(relx=0.5, rely=0.4, anchor="center", width=550, height=350)

        inner = ttk.Frame(card, style="CardInner.TFrame")
        inner.pack(fill="both", expand=True, padx=20, pady=20)

        ttk.Label(inner, text="Generar Reporte de Facturas", font=("Segoe UI", 14, "bold")).pack(pady=(0, 20))

        # Selector de Rango Rápido
        ttk.Label(inner, text="Selecciona el rango de fechas:", style="Muted.TLabel").pack(anchor="w")

        self.var_filtro = tk.StringVar(value="Mes Actual")
        opciones_fecha = ["Hoy", "Ayer", "Mes Actual", "Mes Anterior", "Todo el Histórico"]
        if HAS_CALENDAR:
            opciones_fecha.append("Rango Personalizado")

        self.cmb_filtro = ttk.Combobox(inner, textvariable=self.var_filtro, values=opciones_fecha, state="readonly",
                                       font=("Segoe UI", 11))
        self.cmb_filtro.pack(fill="x", pady=(5, 15))
        self.cmb_filtro.bind("<<ComboboxSelected>>", self._on_filtro_change)

        # Contenedor para Calendarios (Oculto por defecto)
        self.frame_calendarios = ttk.Frame(inner, style="CardInner.TFrame")

        # --- BLINDAJE: Solo dibujamos el calendario si la librería está instalada ---
        if HAS_CALENDAR:
            ttk.Label(self.frame_calendarios, text="Desde:").grid(row=0, column=0, padx=(0, 5), sticky="w")
            self.cal_desde = DateEntry(self.frame_calendarios, width=12, background='darkblue', foreground='white',
                                       borderwidth=2, date_pattern='yyyy-mm-dd')
            self.cal_desde.grid(row=0, column=1, padx=(0, 20), sticky="w")

            ttk.Label(self.frame_calendarios, text="Hasta:").grid(row=0, column=2, padx=(0, 5), sticky="w")
            self.cal_hasta = DateEntry(self.frame_calendarios, width=12, background='darkblue', foreground='white',
                                       borderwidth=2, date_pattern='yyyy-mm-dd')
            self.cal_hasta.grid(row=0, column=3, sticky="w")
        else:
            ttk.Label(self.frame_calendarios,
                      text="⚠️ Librería 'tkcalendar' no detectada. Usa la terminal para instalarla.",
                      foreground="red").grid(row=0, column=0)

        # Botón de Exportar
        btn_exportar = ttk.Button(inner, text="Generar y Guardar Excel", style="Primary.TButton",
                                  command=self._generar_reporte)
        btn_exportar.pack(side="bottom", fill="x", ipady=5, pady=(20, 0))

    def _on_filtro_change(self, event=None):
        if self.var_filtro.get() == "Rango Personalizado":
            self.frame_calendarios.pack(fill="x", pady=(0, 15))
        else:
            self.frame_calendarios.pack_forget()

    def _generar_reporte(self):
        if not HAS_EXCEL:
            messagebox.showerror("Error", "Falta instalar openpyxl. Ejecuta: pip install openpyxl en tu terminal.")
            return

        filtro = self.var_filtro.get()
        now = datetime.now()

        if filtro == "Hoy":
            start = now.replace(hour=0, minute=0, second=0)
            end = now.replace(hour=23, minute=59, second=59)
        elif filtro == "Ayer":
            ayer = now - timedelta(days=1)
            start = ayer.replace(hour=0, minute=0, second=0)
            end = ayer.replace(hour=23, minute=59, second=59)
        elif filtro == "Mes Actual":
            start = now.replace(day=1, hour=0, minute=0, second=0)
            end = now.replace(hour=23, minute=59, second=59)
        elif filtro == "Mes Anterior":
            primero_este_mes = now.replace(day=1)
            ultimo_mes_pasado = primero_este_mes - timedelta(days=1)
            start = ultimo_mes_pasado.replace(day=1, hour=0, minute=0, second=0)
            end = ultimo_mes_pasado.replace(hour=23, minute=59, second=59)
        elif filtro == "Rango Personalizado" and HAS_CALENDAR:
            fecha_inicio = self.cal_desde.get_date()
            fecha_fin = self.cal_hasta.get_date()
            start = datetime.combine(fecha_inicio, datetime.min.time())
            end = datetime.combine(fecha_fin, datetime.max.time())
        else:
            start = None
            end = None

        db = SessionLocal()
        try:
            query = db.query(FacturaGuardada)
            if start:
                query = query.filter(FacturaGuardada.fecha_registro >= start, FacturaGuardada.fecha_registro <= end)

            facturas = query.all()

            if not facturas:
                messagebox.showinfo("Sin Datos", "No se encontraron facturas registradas en el rango seleccionado.")
                return

            fecha_str_archivo = now.strftime("%Y%m%d")
            nombre_filtro = filtro.replace(" ", "_")
            nombre_sugerido = f"REPORTE_FactBot_{nombre_filtro}_{fecha_str_archivo}.xlsx"

            ruta = filedialog.asksaveasfilename(
                initialfile=nombre_sugerido,
                defaultextension=".xlsx",
                filetypes=[("Archivo Excel", "*.xlsx")],
                title="Guardar Reporte"
            )

            if not ruta: return

            # --- GENERACIÓN DEL EXCEL CON OPENPYXL ---
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte de Emisiones"

            encabezados = [
                "ID Interno", "Fecha de Registro", "Archivo Origen", "Hoja",
                "Proveedor", "Sucursal", "Receptor (RFC)", "Uso CFDI",
                "Método Pago", "Forma Pago", "Moneda", "Tipo de Cambio",
                "Subtotal", "Monto Total", "Cant. Conceptos", "Modo de Ejecución",
                "Folio Interno", "Estado Final"
            ]
            ws.append(encabezados)

            for col in range(1, len(encabezados) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for fac in facturas:
                cant_conceptos = 0
                subtotal = 0.0
                try:
                    conceptos_lista = json.loads(fac.conceptos_json)
                    cant_conceptos = len(conceptos_lista)
                    for c in conceptos_lista:
                        cant = float(c.get("cantidad", 0) or 0)
                        precio = float(c.get("precio_unitario", 0) or 0)
                        subtotal += (cant * precio)
                except Exception:
                    pass

                fecha_str = fac.fecha_registro.strftime("%Y-%m-%d %H:%M") if fac.fecha_registro else "Desconocida"
                moneda = "USD" if fac.es_usd else "MXN"
                modo = "Automático" if fac.emitir_y_enviar else "Manual"
                tipo_cambio = fac.tipo_cambio if fac.es_usd and fac.tipo_cambio else "N/A"

                fila = [
                    fac.id, fecha_str, fac.archivo_origen or "-", fac.hoja_origen or "-",
                                       fac.proveedor or "-", fac.sucursal or "-", fac.rfc_cliente or "-",
                                       fac.uso_cfdi or "-", fac.metodo_pago or "-", fac.forma_pago or "-",
                    moneda, tipo_cambio, subtotal, fac.total or 0.0, cant_conceptos,
                    modo, fac.folio_fiscal or "NO ASIGNADO", fac.estado or "-"
                ]
                ws.append(fila)

            for row in range(2, ws.max_row + 1):
                ws[f'M{row}'].number_format = '"$"#,##0.00'
                ws[f'N{row}'].number_format = '"$"#,##0.00'

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = (max_length + 2)

            wb.save(ruta)
            messagebox.showinfo("Éxito",
                                f"Reporte Excel generado correctamente.\n\nSe exportaron {len(facturas)} facturas.")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar el reporte:\n{str(e)}")
        finally:
            db.close()

    def _toggle_theme(self):
        self.controller.toggle_theme()

    def on_theme_changed(self):
        for widget in self.winfo_children():
            widget.destroy()
        self._build_ui()
